import re
from datetime import date, datetime

import openpyxl

from gestor_credito.db.database import ESTADO_EN_ESPERA_CONSTANCIA, ESTADO_EN_PROCESO, get_connection

# Columnas de la bitácora de MIDESA cuyo valor es una fecha y debe normalizarse a
# texto ISO (YYYY-MM-DD) sin importar si openpyxl las entrega como datetime o texto.
DATE_FIELDS = {
    "fecha_registro",
    "fecha_ultima_gestion",
    "proxima_gestion",
    "fecha_envio_siaf",
    "fecha_decision",
}

INT_FIELDS = {"dias_en_gestion"}
FLOAT_FIELDS = {"monto_solicitado"}

_SUFIJO_MANUAL_AUTO = re.compile(r"\(\s*(manual|auto|autom[aá]tico)\s*\)", re.IGNORECASE)
_ESPACIOS_ALREDEDOR_DE_BARRA = re.compile(r"\s*/\s*")
_ESPACIOS_MULTIPLES = re.compile(r"\s+")

# Encabezado del Excel (normalizado) -> columna interna. La plantilla real de MIDESA
# ("MachoteBaseDeDatos.xlsx", hoja 01_Bitacora_Piloto) trae los encabezados con un
# salto de línea interno y un sufijo "(Manual)"/"(Auto)" pegado al nombre del campo
# (p. ej. "ID Caso\n(Auto)"); _normalize_header() quita eso antes de buscar aquí. Los
# VALORES de las celdas no se tocan, solo los encabezados.
COLUMN_ALIASES = {
    "id caso": "id_caso",
    "fecha registro": "fecha_registro",
    "canal/origen": "canal_origen",
    "no. presolicitud": "no_presolicitud",
    "ejecutivo": "ejecutivo",
    "constancia solicitada": "constancia_solicitada",
    "empresa convenio": "empresa_convenio",
    "nombre del cliente": "nombre",
    "identificación": "cedula",
    "identificacion": "cedula",
    "teléfono": "telefono",
    "telefono": "telefono",
    "monto solicitado": "monto_solicitado",
    "destino del crédito": "destino_credito",
    "destino del credito": "destino_credito",
    "microseguro": "microseguro",
    "estado solicitud": "estado_solicitud",
    "etapa proceso": "etapa_proceso",
    "responsable actual": "responsable_actual",
    "fecha última gestión": "fecha_ultima_gestion",
    "fecha ultima gestion": "fecha_ultima_gestion",
    "próxima gestión": "proxima_gestion",
    "proxima gestion": "proxima_gestion",
    "días en gestión": "dias_en_gestion",
    "dias en gestion": "dias_en_gestion",
    "alerta seguimiento": "alerta_seguimiento",
    "¿requiere registro/acción siaf?": "requiere_siaf",
    "¿requiere registro/accion siaf?": "requiere_siaf",
    "fecha envío siaf": "fecha_envio_siaf",
    "fecha envio siaf": "fecha_envio_siaf",
    "fecha decisión": "fecha_decision",
    "fecha decision": "fecha_decision",
    "decisión": "decision",
    "decision": "decision",
    "motivo no aplica/desistimiento": "motivo_no_aplica",
    "observaciones": "observaciones",
}

CASO_COLUMNS = [
    "id_caso", "no_presolicitud", "fecha_registro", "canal_origen", "ejecutivo",
    "empresa_convenio", "monto_solicitado", "destino_credito", "microseguro",
    "estado_solicitud", "etapa_proceso", "responsable_actual", "fecha_ultima_gestion",
    "proxima_gestion", "dias_en_gestion", "alerta_seguimiento", "requiere_siaf",
    "fecha_envio_siaf", "fecha_decision", "decision", "motivo_no_aplica",
    "observaciones", "constancia_solicitada",
]


class ImportSummary:
    def __init__(self):
        self.clientes_nuevos = 0
        self.casos_nuevos = 0
        self.casos_actualizados = 0
        self.filas_omitidas = []

    def __repr__(self):
        return (
            f"ImportSummary(clientes_nuevos={self.clientes_nuevos}, "
            f"casos_nuevos={self.casos_nuevos}, casos_actualizados={self.casos_actualizados}, "
            f"filas_omitidas={len(self.filas_omitidas)})"
        )


def import_bitacora(file_path):
    summary = ImportSummary()

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = _read_headers(sheet)

    faltantes = {"cedula", "nombre"} - headers.keys()
    if faltantes:
        raise ValueError(
            f"Faltan columnas obligatorias en el Excel: {', '.join(sorted(faltantes))}"
        )

    conn = get_connection()
    try:
        for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            data = _row_to_dict(row, headers)
            cedula = (data.get("cedula") or "").strip()
            clave_caso = (data.get("no_presolicitud") or data.get("id_caso") or "").strip()

            if not cedula or not clave_caso:
                summary.filas_omitidas.append((row_number, "falta cédula o No. Presolicitud/ID Caso"))
                continue

            cliente_id, cliente_nuevo = _upsert_cliente(conn, cedula, data)
            if cliente_nuevo:
                summary.clientes_nuevos += 1

            caso_nuevo = _upsert_caso(conn, cliente_id, clave_caso, data)
            if caso_nuevo:
                summary.casos_nuevos += 1
            else:
                summary.casos_actualizados += 1

        conn.commit()
    finally:
        conn.close()

    return summary


def _read_headers(sheet):
    headers = {}
    primera_fila = next(sheet.iter_rows(min_row=1, max_row=1))
    for col_index, cell in enumerate(primera_fila):
        field = COLUMN_ALIASES.get(_normalize_header(cell.value))
        if field:
            headers[field] = col_index
    return headers


def _normalize_header(value):
    text = str(value or "").replace("\n", " ").replace("\r", " ")
    text = _SUFIJO_MANUAL_AUTO.sub("", text)
    text = _ESPACIOS_ALREDEDOR_DE_BARRA.sub("/", text)
    text = _ESPACIOS_MULTIPLES.sub(" ", text)
    return text.strip().lower()


def _row_to_dict(row, headers):
    data = {}
    for field, col_index in headers.items():
        value = row[col_index].value if col_index < len(row) else None

        if field in DATE_FIELDS:
            value = _normalize_date(value)
        elif field in INT_FIELDS:
            value = _to_int(value)
        elif field in FLOAT_FIELDS:
            value = _to_float(value)
        elif value is not None:
            # Todo lo demás se guarda como texto, incluso si Excel entregó un
            # número (p. ej. "No. Presolicitud" puede venir como int 2877).
            value = str(value).strip()

        data[field] = value
    return data


def _normalize_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _to_int(value):
    if value in (None, ""):
        return None
    return int(value)


def _to_float(value):
    if value in (None, ""):
        return None
    return float(value)


def _upsert_cliente(conn, cedula, data):
    row = conn.execute(
        "SELECT id, nombre, telefono FROM cliente WHERE cedula = ?", (cedula,)
    ).fetchone()
    nombre = data.get("nombre") or ""
    telefono = data.get("telefono")

    if row is None:
        cur = conn.execute(
            "INSERT INTO cliente (cedula, nombre, telefono) VALUES (?, ?, ?)",
            (cedula, nombre, telefono),
        )
        return cur.lastrowid, True

    cliente_id, nombre_actual, telefono_actual = row
    if nombre != nombre_actual or telefono != telefono_actual:
        conn.execute(
            "UPDATE cliente SET nombre = ?, telefono = ?, fecha_actualizacion = datetime('now') "
            "WHERE id = ?",
            (nombre, telefono, cliente_id),
        )
    return cliente_id, False


def _upsert_caso(conn, cliente_id, clave_caso, data):
    existing = conn.execute(
        "SELECT id, estado_solicitud FROM caso WHERE cliente_id = ? AND clave_caso = ?",
        (cliente_id, clave_caso),
    ).fetchone()

    valores = {column: data.get(column) for column in CASO_COLUMNS}

    if existing is None:
        columnas = ["cliente_id", "clave_caso", *CASO_COLUMNS, "origen_ultima_modificacion"]
        placeholders = ", ".join("?" for _ in columnas)
        params = [cliente_id, clave_caso, *valores.values(), "excel"]
        conn.execute(f"INSERT INTO caso ({', '.join(columnas)}) VALUES ({placeholders})", params)
        return True

    caso_id, estado_anterior = existing
    estado_nuevo = valores["estado_solicitud"]

    set_sql = [f"{column} = ?" for column in CASO_COLUMNS]
    params = list(valores.values())

    set_sql.append("origen_ultima_modificacion = ?")
    params.append("excel")
    set_sql.append("fecha_actualizacion_registro = datetime('now')")

    if estado_nuevo != estado_anterior:
        set_sql.append("estado_solicitud_fecha_cambio = datetime('now')")
        # Solo esta transición puntual arranca el reloj de 48h de la Alerta
        # "Constancia en mano": no cualquier salida de "En espera de
        # constancia" (p. ej. a "No aplica" o "Cliente desistió" no cuenta).
        if estado_anterior == ESTADO_EN_ESPERA_CONSTANCIA and estado_nuevo == ESTADO_EN_PROCESO:
            set_sql.append("constancia_recibida_fecha = datetime('now')")

    params.append(caso_id)
    conn.execute(f"UPDATE caso SET {', '.join(set_sql)} WHERE id = ?", params)
    return False
