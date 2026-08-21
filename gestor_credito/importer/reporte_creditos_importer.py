import re
import sqlite3
from datetime import date, datetime

import openpyxl

from gestor_credito.db.database import get_connection

DATE_FIELDS = {"fecha_desembolso", "fecha_vencimiento"}
INT_FIELDS = {"plazo_credito", "numero_cuotas", "cuotas_pagadas"}
FLOAT_FIELDS = {"monto_desembolsado", "saldo_principal", "saldo_intereses"}

_ESPACIOS_MULTIPLES = re.compile(r"\s+")

# Encabezado del Excel (normalizado) -> columna interna. Los encabezados reales
# de recursos/reporte.xlsx ("REPORTE DE DATOS DE CREDITOS") usan guion bajo
# (p. ej. "NO_IDENTIFICACION"), sin saltos de línea ni sufijos — mucho más
# limpios que la bitácora de MIDESA. _normalize_header() igual convierte "_" a
# espacio antes de buscar acá, para tolerar tanto "NO_CREDITO" como una futura
# variante con espacios ("NO CREDITO"), y se incluyen también las formas con
# preposición que usó el usuario al pedir el módulo ("NOMBRE del CLIENTE",
# "PLAZO del CREDITO") por si un reporte futuro trae esos encabezados en vez
# de los que tiene el archivo real. Los VALORES de las celdas no se tocan,
# solo los encabezados.
COLUMN_ALIASES = {
    "fecha desembolso": "fecha_desembolso",
    "fecha vencimiento": "fecha_vencimiento",
    "no credito": "no_credito",
    "monto desembolsado": "monto_desembolsado",
    "nombre cliente": "nombre_cliente",
    "nombre del cliente": "nombre_cliente",
    "estado credito": "estado_credito",
    "empresa de convenio": "empresa_convenio",
    "empresa convenio": "empresa_convenio",
    "no identificacion": "cedula",
    "plazo credito": "plazo_credito",
    "plazo del credito": "plazo_credito",
    "numero cuotas": "numero_cuotas",
    "numero de cuotas": "numero_cuotas",
    "cuotas pagadas": "cuotas_pagadas",
    # Agregadas 2026-08-21, pedido explícito del usuario: de las 13 columnas
    # que resaltó en el reporte real, estas dos son las únicas dos genuinamente
    # nuevas que sí hace falta importar — alimentan la columna calculada
    # "Saldo a la fecha" en Historial de Créditos (ver creditos_panel.py), no
    # tienen columna propia visible. ES_CONVENIO y FECHA_REPORTE, también
    # resaltadas, se descartaron a propósito (pedido explícito del usuario:
    # "ignórala ya que no la vamos a usar").
    "saldo principal": "saldo_principal",
    "saldo intereses": "saldo_intereses",
}

CREDITO_COLUMNS = [
    "cedula", "nombre_cliente", "fecha_desembolso", "fecha_vencimiento",
    "monto_desembolsado", "estado_credito", "empresa_convenio",
    "plazo_credito", "numero_cuotas", "cuotas_pagadas",
    "saldo_principal", "saldo_intereses",
]


class ImportResumenCreditos:
    def __init__(self):
        self.creditos_nuevos = 0
        self.creditos_actualizados = 0
        self.filas_omitidas = []

    def __repr__(self):
        return (
            f"ImportResumenCreditos(creditos_nuevos={self.creditos_nuevos}, "
            f"creditos_actualizados={self.creditos_actualizados}, "
            f"filas_omitidas={len(self.filas_omitidas)})"
        )


def import_reporte_creditos(file_path):
    """Lee recursos/reporte.xlsx (o .xls/.xlsm — cualquier archivo que
    openpyxl pueda abrir) y hace upsert de cada fila contra reporte_credito,
    igual que import_bitacora() hace con caso: no_credito es la clave real,
    así que una reimportación periódica actualiza la fila existente en vez de
    duplicarla. Nunca borra filas que ya no aparezcan en un reimport
    posterior — mismo criterio que la bitácora."""
    resumen = ImportResumenCreditos()

    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook.active
    fila_encabezado, headers = _localizar_fila_encabezado(sheet)

    faltantes = {"no_credito", "cedula"} - headers.keys()
    if faltantes:
        raise ValueError(
            f"Faltan columnas obligatorias en el Excel: {', '.join(sorted(faltantes))}"
        )

    conn = get_connection()
    try:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=fila_encabezado + 1), start=fila_encabezado + 1
        ):
            try:
                data = _row_to_dict(row, headers)
                no_credito = (data.get("no_credito") or "").strip()
                cedula = (data.get("cedula") or "").strip()
                nombre_cliente = (data.get("nombre_cliente") or "").strip()

                if not no_credito or not cedula or not nombre_cliente:
                    resumen.filas_omitidas.append(
                        (row_number, "falta No. Crédito, Cédula o Nombre del Cliente")
                    )
                    continue

                nuevo = _upsert_credito(conn, no_credito, data)
                if nuevo:
                    resumen.creditos_nuevos += 1
                else:
                    resumen.creditos_actualizados += 1
            except (ValueError, TypeError, sqlite3.Error) as error:
                # Una fila con un dato inválido (p. ej. texto no numérico en
                # PLAZO_CREDITO/CUOTAS_PAGADAS/MONTO_DESEMBOLSADO) no debe
                # perder TODO el lote: antes, una excepción sin capturar acá
                # se propagaba hasta afuera del bucle, y como conn.commit()
                # solo ocurre una vez al final, ninguna fila ya procesada en
                # esta importación quedaba guardada. Se omite solo esta fila,
                # igual que el caso de datos faltantes arriba, y se sigue con
                # las demás (reporte real del usuario, 2026-08-16).
                resumen.filas_omitidas.append((row_number, f"dato inválido: {error}"))

        conn.commit()
    finally:
        conn.close()

    return resumen


# Cuántas filas se revisan buscando el encabezado real — el reporte real
# trae una fila de título antes de los encabezados (bug real encontrado
# 2026-08-21 validando contra un archivo real de producción: "REPORTE DE
# DATOS DE CREDITOS AL DD/MM/AAAA" en la fila 2, con la fila 1 completamente
# vacía, y los encabezados reales recién en la fila 3 — con el código
# anterior, que asumía la fila 1 a secas, el import fallaba siempre con
# "Faltan columnas obligatorias en el Excel: cedula, no_credito" contra
# cualquier archivo con este formato real). 15 filas es margen de sobra para
# cualquier preámbulo razonable sin arriesgar confundir una fila de datos
# con el encabezado.
_MAX_FILAS_BUSQUEDA_ENCABEZADO = 15


def _localizar_fila_encabezado(sheet):
    """Devuelve (numero_fila, headers) de la fila, entre las primeras
    _MAX_FILAS_BUSQUEDA_ENCABEZADO, que tiene más coincidencias contra
    COLUMN_ALIASES — no se puede asumir que los encabezados siempre están en
    la fila 1 (ver comentario de la constante arriba). Una fila de título o
    una fila de datos normalmente no coincide con ningún alias conocido (0
    coincidencias), así que la fila de encabezados reales gana por amplio
    margen incluso sin conocer de antemano en qué fila está.

    Si ninguna fila tiene coincidencias, devuelve (1, {}) — el llamador ya
    valida que falten "no_credito"/"cedula" y lanza el mismo error de
    siempre, con un mensaje que sigue siendo correcto aunque la causa real
    sea "no se encontró ninguna fila de encabezados", no solo "esta fila no
    los tiene"."""
    mejor_numero_fila = 1
    mejor_headers = {}
    mejor_cantidad = 0

    for numero_fila, fila in enumerate(
        sheet.iter_rows(min_row=1, max_row=_MAX_FILAS_BUSQUEDA_ENCABEZADO), start=1
    ):
        headers_fila = {}
        for col_index, cell in enumerate(fila):
            field = COLUMN_ALIASES.get(_normalize_header(cell.value))
            if field:
                headers_fila[field] = col_index
        if len(headers_fila) > mejor_cantidad:
            mejor_cantidad = len(headers_fila)
            mejor_numero_fila = numero_fila
            mejor_headers = headers_fila

    return mejor_numero_fila, mejor_headers


def _normalize_header(value):
    text = str(value or "").replace("_", " ").replace("\n", " ").replace("\r", " ")
    text = _ESPACIOS_MULTIPLES.sub(" ", text)
    return text.strip().lower()


def _row_to_dict(row, headers):
    data = {}
    for field, col_index in headers.items():
        value = row[col_index].value if col_index < len(row) else None

        if field in DATE_FIELDS:
            value = _normalize_date(value)
        elif field in INT_FIELDS:
            value = _to_int(value)
        elif field in FLOAT_FIELDS:
            value = _to_float(value)
        elif value is not None:
            # Todo lo demás se guarda como texto, incluso si Excel entregó un
            # número (mismo motivo que en excel_importer.py: NO_CREDITO/
            # NO_IDENTIFICACION podrían venir como número en algún reporte).
            value = str(value).strip()

        data[field] = value
    return data


def _normalize_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _to_int(value):
    if value in (None, ""):
        return None
    return int(value)


def _to_float(value):
    if value in (None, ""):
        return None
    return float(value)


def _upsert_credito(conn, no_credito, data):
    existing = conn.execute(
        "SELECT id, estado_credito FROM reporte_credito WHERE no_credito = ?", (no_credito,)
    ).fetchone()

    if existing is None and no_credito.isdigit():
        # El mismo crédito puede llegar con distinto formato de texto entre
        # reimportaciones de reporte.xlsx: si Excel entrega la celda NO_CREDITO
        # como número en vez de texto (ver _row_to_dict), "0012456" de un
        # reporte se vuelve "12456" en el siguiente. La comparación exacta de
        # arriba no encuentra la fila ya existente y el upsert insertaba un
        # crédito duplicado — "Historial de Créditos" luego mostraba el mismo
        # crédito dos veces para el mismo cliente (reporte real del usuario,
        # 2026-08-16). CAST a INTEGER compara el valor numérico ignorando
        # ceros a la izquierda; no_credito.isdigit() evita este camino si
        # algún no_credito futuro trae letras (no ocurre en los datos reales
        # verificados, pero CAST silenciosamente da 0 con texto no numérico).
        # No se sobreescribe no_credito en el UPDATE de más abajo, así que el
        # primer formato de texto importado para un crédito es el que queda
        # guardado permanentemente como su identidad.
        existing = conn.execute(
            "SELECT id, estado_credito FROM reporte_credito "
            "WHERE no_credito != ? AND CAST(no_credito AS INTEGER) = CAST(? AS INTEGER)",
            (no_credito, no_credito),
        ).fetchone()

    valores = {column: data.get(column) for column in CREDITO_COLUMNS}

    if existing is None:
        columnas = ["no_credito", *CREDITO_COLUMNS]
        placeholders = ", ".join("?" for _ in columnas)
        params = [no_credito, *valores.values()]
        conn.execute(
            f"INSERT INTO reporte_credito ({', '.join(columnas)}, estado_credito_fecha_cambio) "
            f"VALUES ({placeholders}, datetime('now'))",
            params,
        )
        return True

    credito_id, estado_anterior = existing
    estado_nuevo = valores["estado_credito"]

    set_sql = [f"{column} = ?" for column in CREDITO_COLUMNS]
    set_sql.append("fecha_actualizacion_registro = datetime('now')")
    # Mismo patrón que estado_solicitud_fecha_cambio en caso (excel_importer.py):
    # solo se pisa a "ahora" cuando estado_credito realmente cambia, nunca en
    # cada reimport que no cambia nada — así la vista "Finalizados" ordena por
    # cuándo se detectó el cierre real, no por cuándo se volvió a importar el
    # mismo reporte sin novedades.
    if estado_nuevo != estado_anterior:
        set_sql.append("estado_credito_fecha_cambio = datetime('now')")
    params = [*valores.values(), credito_id]
    conn.execute(f"UPDATE reporte_credito SET {', '.join(set_sql)} WHERE id = ?", params)
    return False
